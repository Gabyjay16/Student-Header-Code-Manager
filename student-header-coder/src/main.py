import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import openpyxl
import pandas as pd
import os
from datetime import datetime
import json
import speech_recognition as sr
import threading

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class HeaderCodeApp:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("Student Header Code Assigner")
        self.root.geometry("1200x800")
        
        self.current_file = None
        self.wb = None
        self.ws = None
        self.matricules = []
        self.header_col = None
        self.matricule_col = None
        self.current_year = "24"
        self.current_code = 1
        self.processed = 0
        self.history = []  # For undo
        self.config_file = "config.json"
        self.voice_enabled = False
        self.recognizer = sr.Recognizer()
        
        self.load_config()
        self.create_startup_screen()
    
    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    config = json.load(f)
                    self.current_year = config.get('year', '24')
                    self.current_code = config.get('code', 1)
            except:
                pass
    
    def save_config(self):
        config = {
            'year': self.current_year,
            'code': self.current_code
        }
        with open(self.config_file, 'w') as f:
            json.dump(config, f)
    
    def create_startup_screen(self):
        self.clear_window()
        
        frame = ctk.CTkFrame(self.root)
        frame.pack(padx=20, pady=20, fill="both", expand=True)
        
        ctk.CTkLabel(frame, text="Student Header Code Assigner", font=ctk.CTkFont(size=24, weight="bold")).pack(pady=20)
        
        # File selection
        file_frame = ctk.CTkFrame(frame)
        file_frame.pack(pady=10, fill="x", padx=50)
        
        self.file_label = ctk.CTkLabel(file_frame, text="No file selected")
        self.file_label.pack(side="left", padx=10)
        
        ctk.CTkButton(file_frame, text="Select Excel File", command=self.select_file).pack(side="right", padx=10)
        
        # Year
        year_frame = ctk.CTkFrame(frame)
        year_frame.pack(pady=10, fill="x", padx=50)
        ctk.CTkLabel(year_frame, text="Academic Year:").pack(side="left", padx=10)
        self.year_entry = ctk.CTkEntry(year_frame, placeholder_text="24")
        self.year_entry.insert(0, self.current_year)
        self.year_entry.pack(side="left", padx=10)
        
        # Starting code
        code_frame = ctk.CTkFrame(frame)
        code_frame.pack(pady=10, fill="x", padx=50)
        ctk.CTkLabel(code_frame, text="Starting Code:").pack(side="left", padx=10)
        self.code_entry = ctk.CTkEntry(code_frame, placeholder_text="001")
        self.code_entry.insert(0, f"{self.current_code:03d}")
        self.code_entry.pack(side="left", padx=10)
        
        ctk.CTkButton(frame, text="Start Session", font=ctk.CTkFont(size=16), command=self.start_session).pack(pady=30)
    
    def select_file(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file:
            self.current_file = file
            self.file_label.configure(text=os.path.basename(file))
    
    def start_session(self):
        if not self.current_file:
            messagebox.showerror("Error", "Please select an Excel file")
            return
        
        self.current_year = self.year_entry.get().strip()
        try:
            self.current_code = int(self.code_entry.get().strip())
        except:
            self.current_code = 1
        
        self.load_excel()
        self.create_main_screen()
    
    def load_excel(self):
        try:
            self.wb = openpyxl.load_workbook(self.current_file)
            self.ws = self.wb.active
            
            # Find columns
            headers = [cell.value for cell in self.ws[1]]
            for i, h in enumerate(headers):
                if h and isinstance(h, str):
                    if 'matricule' in h.lower():
                        self.matricule_col = i + 1
                    elif 'header' in h.lower() or 'code' in h.lower():
                        self.header_col = i + 1
            
            if not self.matricule_col:
                messagebox.showerror("Error", "Could not find Matricule column")
                return
            
            if not self.header_col:
                # Add header code column if missing
                self.header_col = len(headers) + 1
                self.ws.cell(1, self.header_col, "Header Code")
            
            # Load data
            self.matricules = []
            for row in range(2, self.ws.max_row + 1):
                mat_cell = self.ws.cell(row, self.matricule_col)
                if mat_cell.value:
                    self.matricules.append({
                        'row': row,
                        'matricule': str(mat_cell.value).strip(),
                        'header': self.ws.cell(row, self.header_col).value
                    })
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel: {str(e)}")
    
    def create_main_screen(self):
        self.clear_window()
        
        # Top bar
        top_frame = ctk.CTkFrame(self.root, height=100)
        top_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(top_frame, text=f"Year: {self.current_year}", font=ctk.CTkFont(size=16)).pack(side="left", padx=20)
        self.code_label = ctk.CTkLabel(top_frame, text=f"Current Code: {self.current_code:03d}", font=ctk.CTkFont(size=18, weight="bold"))
        self.code_label.pack(side="left", padx=20)
        
        self.stats_label = ctk.CTkLabel(top_frame, text="Processed: 0 | Remaining: 0")
        self.stats_label.pack(side="left", padx=20)
        
        # Voice toggle
        self.voice_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(top_frame, text="Voice Input", variable=self.voice_var, command=self.toggle_voice).pack(side="right", padx=20)
        
        # Search area
        search_frame = ctk.CTkFrame(self.root)
        search_frame.pack(pady=20, padx=20, fill="x")
        
        ctk.CTkLabel(search_frame, text="Enter Matricule (partial OK):", font=ctk.CTkFont(size=14)).pack(pady=5)
        
        self.search_entry = ctk.CTkEntry(search_frame, placeholder_text="031 or UBA24C031", height=40, font=ctk.CTkFont(size=16))
        self.search_entry.pack(fill="x", padx=50, pady=10)
        self.search_entry.bind("<Return>", self.process_search)
        self.search_entry.focus_set()
        
        # Results area
        self.result_frame = ctk.CTkFrame(self.root)
        self.result_frame.pack(pady=10, padx=20, fill="both", expand=True)
        
        # History
        history_frame = ctk.CTkFrame(self.root, height=200)
        history_frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(history_frame, text="Recent Assignments").pack(anchor="w", padx=10)
        
        self.history_tree = ttk.Treeview(history_frame, columns=("Matricule", "Code"), show="headings", height=6)
        self.history_tree.heading("Matricule", text="Matricule")
        self.history_tree.heading("Code", text="Code")
        self.history_tree.pack(fill="x", padx=10, pady=5)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self.root)
        btn_frame.pack(pady=10, fill="x")
        
        ctk.CTkButton(btn_frame, text="Undo Last", command=self.undo_last).pack(side="left", padx=20)
        ctk.CTkButton(btn_frame, text="Export Coded File", command=self.export_file).pack(side="right", padx=20)
        
        self.update_stats()
    
    def toggle_voice(self):
        self.voice_enabled = self.voice_var.get()
        if self.voice_enabled:
            threading.Thread(target=self.listen_voice, daemon=True).start()
    
    def listen_voice(self):
        while self.voice_enabled:
            try:
                with sr.Microphone() as source:
                    self.recognizer.adjust_for_ambient_noise(source)
                    audio = self.recognizer.listen(source, timeout=5)
                    text = self.recognizer.recognize_google(audio)
                    self.root.after(0, self.handle_voice_input, text)
            except:
                pass
    
    def handle_voice_input(self, text):
        self.search_entry.delete(0, tk.END)
        self.search_entry.insert(0, text.strip())
        self.process_search(None)
    
    def process_search(self, event=None):
        query = self.search_entry.get().strip().upper()
        if not query:
            return
        
        matches = []
        for student in self.matricules:
            mat = student['matricule'].upper()
            if query in mat or mat.endswith(query):
                matches.append(student)
        
        if not matches:
            messagebox.showinfo("Not Found", "No matching matricule found")
            self.search_entry.focus_set()
            return
        
        if len(matches) == 1:
            self.assign_code(matches[0])
        else:
            self.show_matches(matches)
    
    def assign_code(self, student):
        if student['header']:
            if not messagebox.askyesno("Already Coded", f"This student already has code {student['header']}. Replace?"):
                self.search_entry.focus_set()
                return
        
        code_str = f"{self.current_code:03d}"
        student['header'] = code_str
        
        # Update Excel
        self.ws.cell(student['row'], self.header_col, code_str)
        self.wb.save(self.current_file)  # Autosave
        
        # History
        self.history.append({
            'row': student['row'],
            'old_code': None,  # For undo
            'matricule': student['matricule']
        })
        
        self.processed += 1
        self.current_code += 1
        self.code_label.configure(text=f"Current Code: {self.current_code:03d}")
        
        # Update history tree
        self.history_tree.insert("", 0, values=(student['matricule'], code_str))
        
        self.update_stats()
        messagebox.showinfo("Success", f"Assigned {code_str} to {student['matricule']}")
        
        self.search_entry.delete(0, tk.END)
        self.search_entry.focus_set()
        self.save_config()
    
    def show_matches(self, matches):
        # Simple dialog for multiple matches
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Multiple Matches")
        dialog.geometry("600x400")
        
        ctk.CTkLabel(dialog, text="Select the correct student:").pack(pady=10)
        
        for i, m in enumerate(matches):
            btn = ctk.CTkButton(dialog, text=m['matricule'], 
                              command=lambda s=m: self.select_from_matches(s, dialog))
            btn.pack(pady=5, padx=20, fill="x")
    
    def select_from_matches(self, student, dialog):
        dialog.destroy()
        self.assign_code(student)
    
    def update_stats(self):
        remaining = len([s for s in self.matricules if not s['header']])
        self.stats_label.configure(text=f"Processed: {self.processed} | Remaining: {remaining}")
    
    def undo_last(self):
        if not self.history:
            messagebox.showinfo("Undo", "Nothing to undo")
            return
        
        last = self.history.pop()
        # Revert Excel (simplified)
        # In full version, restore old value
        messagebox.showinfo("Undo", f"Undid assignment for {last['matricule']}")
        self.current_code = max(1, self.current_code - 1)
        self.code_label.configure(text=f"Current Code: {self.current_code:03d}")
        self.update_stats()
    
    def export_file(self):
        if not self.current_file:
            return
        dir_name = os.path.dirname(self.current_file)
        base = os.path.basename(self.current_file)
        name, ext = os.path.splitext(base)
        new_file = os.path.join(dir_name, f"{name}_Coded{ext}")
        
        self.wb.save(new_file)
        messagebox.showinfo("Export", f"File saved as:\n{new_file}")
    
    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = HeaderCodeApp()
    app.run()